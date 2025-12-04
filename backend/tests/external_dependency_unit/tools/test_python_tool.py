"""
External dependency unit tests for Python tool.

These tests run against a real Code Interpreter service (no mocking of the service).
They verify code execution, error handling, timeout behavior, and file generation.

Requirements:
- CODE_INTERPRETER_BASE_URL must be configured and point to a running service
- Tests use minimal mocking - only mock run_context infrastructure and db lookups
- File store operations execute for real (files are saved and read back)
"""

import io
from datetime import datetime
from unittest.mock import Mock

import pytest
from openpyxl import load_workbook
from pydantic import TypeAdapter
from sqlalchemy.orm import Session

from onyx.chat.emitter import Emitter
from onyx.configs.app_configs import CODE_INTERPRETER_BASE_URL
from onyx.file_store.models import ChatFileType
from onyx.file_store.models import InMemoryChatFile
from onyx.file_store.utils import get_default_file_store
from onyx.tools.models import ChatFile
from onyx.tools.models import (
    LlmPythonExecutionResult,
)
from onyx.tools.models import PythonToolOverrideKwargs
from onyx.tools.tool_implementations.python.code_interpreter_client import (
    CodeInterpreterClient,
)
from onyx.tools.tool_implementations.python.python_tool import PythonTool


# Apply initialize_file_store fixture to all tests in this module
pytestmark = [
    pytest.mark.usefixtures("initialize_file_store"),
    pytest.mark.skip(reason="Tests need Code Interpreter service - skipped for now"),
]


@pytest.fixture
def mock_emitter() -> Emitter:
    """Create a mock emitter for testing."""
    emitter = Mock(spec=Emitter)
    emitter.emit = Mock()
    return emitter


@pytest.fixture
def python_tool(mock_emitter: Emitter) -> PythonTool:
    """Create a PythonTool instance for testing."""
    return PythonTool(tool_id=1, emitter=mock_emitter)


@pytest.fixture
def code_interpreter_client() -> CodeInterpreterClient:
    """Create a real Code Interpreter client for testing."""
    if not CODE_INTERPRETER_BASE_URL:
        pytest.skip("CODE_INTERPRETER_BASE_URL not configured")
    return CodeInterpreterClient()


def test_python_execution_basic(
    python_tool: PythonTool,
    mock_emitter: Emitter,
) -> None:
    """Test basic Python execution with simple code."""
    code = 'print("Hello, World!")'

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert "Hello, World!" in execution_result.stdout
    assert execution_result.stderr == ""
    assert execution_result.exit_code == 0
    assert not execution_result.timed_out
    assert len(execution_result.generated_files) == 0


def test_python_execution_with_syntax_error(
    python_tool: PythonTool,
    mock_emitter: Emitter,
) -> None:
    """Test Python execution with syntax error."""
    code = "print('missing closing quote"

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify error result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.stdout == ""
    assert len(execution_result.stderr) > 0
    assert (
        "SyntaxError" in execution_result.stderr
        or "unterminated" in execution_result.stderr.lower()
    )
    assert execution_result.exit_code != 0
    assert not execution_result.timed_out
    assert execution_result.error is not None or len(execution_result.stderr) > 0
    assert len(execution_result.generated_files) == 0


def test_python_execution_with_runtime_error(
    python_tool: PythonTool,
    mock_emitter: Emitter,
) -> None:
    """Test Python execution with runtime error."""
    code = """
x = 10
y = 0
result = x / y  # Division by zero
print(result)
"""

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify error result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code != 0
    assert (
        "ZeroDivisionError" in execution_result.stderr
        or "division" in execution_result.stderr.lower()
    )
    assert execution_result.error is not None or len(execution_result.stderr) > 0


def test_python_execution_timeout(
    mock_emitter: Emitter,
) -> None:
    """Test execution timeout handling."""
    # Code that will run longer than the timeout
    code = """
import time
time.sleep(10)
print("Should not reach here")
"""

    if not CODE_INTERPRETER_BASE_URL:
        pytest.skip("CODE_INTERPRETER_BASE_URL not configured")

    from unittest.mock import patch

    # Mock the config to use a short timeout
    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CODE_INTERPRETER_DEFAULT_TIMEOUT_MS",
        1000,
    ):
        python_tool = PythonTool(tool_id=1, emitter=mock_emitter)
        override_kwargs = PythonToolOverrideKwargs(chat_files=[])
        result = python_tool.run(
            turn_index=0,
            override_kwargs=override_kwargs,
            code=code,
        )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify timeout result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.timed_out


def test_python_execution_file_generation(
    python_tool: PythonTool,
    mock_emitter: Emitter,
    db_session: Session,  # Needed to initialize DB engine for file_store
) -> None:
    """Test file generation and retrieval."""
    code = """
import csv

# Create a CSV file
with open('test_output.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Name', 'Age', 'City'])
    writer.writerow(['Alice', '30', 'New York'])
    writer.writerow(['Bob', '25', 'San Francisco'])

print("CSV file created successfully")
"""

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == 0
    assert "CSV file created successfully" in execution_result.stdout
    assert len(execution_result.generated_files) == 1

    # Verify file metadata
    generated_file = execution_result.generated_files[0]
    assert generated_file.filename == "test_output.csv"
    assert generated_file.file_link  # File link exists
    assert generated_file.file_link.startswith("http://localhost:3000/api/chat/file/")

    # Extract file_id from file_link
    file_id = generated_file.file_link.split("/")[-1]

    # Verify we can read the file back from the file store
    file_store = get_default_file_store()
    file_io = file_store.read_file(file_id)
    file_content = file_io.read()

    # Verify file content
    assert b"Name,Age,City" in file_content
    assert b"Alice,30,New York" in file_content
    assert b"Bob,25,San Francisco" in file_content


def test_python_execution_with_matplotlib(
    python_tool: PythonTool,
    mock_emitter: Emitter,
    db_session: Session,  # Needed to initialize DB engine for file_store
) -> None:
    """Test matplotlib plot generation."""
    code = """
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import numpy as np

# Generate data
x = np.linspace(0, 10, 100)
y = np.sin(x)

# Create plot
plt.figure(figsize=(10, 6))
plt.plot(x, y)
plt.title('Sine Wave')
plt.xlabel('x')
plt.ylabel('sin(x)')
plt.grid(True)

# Save plot
plt.savefig('sine_wave.png')
print("Plot saved successfully")
"""

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == 0
    assert "Plot saved successfully" in execution_result.stdout
    assert len(execution_result.generated_files) == 1

    # Verify file metadata
    generated_file = execution_result.generated_files[0]
    assert generated_file.filename == "sine_wave.png"
    assert ".png" in generated_file.filename

    # Extract file_id from file_link
    file_id = generated_file.file_link.split("/")[-1]

    # Verify we can read the file back from the file store
    file_store = get_default_file_store()
    file_io = file_store.read_file(file_id)
    file_content = file_io.read()

    # Verify the file is a valid PNG (check PNG magic bytes)
    # PNG magic bytes: 89 50 4E 47 0D 0A 1A 0A
    assert file_content[:8] == b"\x89PNG\r\n\x1a\n"
    assert len(file_content) > 1000  # PNG should be substantial


def test_python_tool_availability_with_url_set(db_session: Session) -> None:
    """Test PythonTool.is_available() returns True when URL is configured."""
    from unittest.mock import patch

    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CODE_INTERPRETER_BASE_URL",
        "http://localhost:8000",
    ):
        assert PythonTool.is_available(db_session) is True


def test_python_tool_availability_without_url(db_session: Session) -> None:
    """Test PythonTool.is_available() returns False when URL is not configured."""
    from unittest.mock import patch

    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CODE_INTERPRETER_BASE_URL",
        None,
    ):
        assert PythonTool.is_available(db_session) is False

    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CODE_INTERPRETER_BASE_URL",
        "",
    ):
        assert PythonTool.is_available(db_session) is False


def test_python_execution_output_truncation(
    mock_emitter: Emitter,
) -> None:
    """Test that large outputs are properly truncated."""
    from unittest.mock import patch

    # Generate code that produces output larger than truncation limit
    code = """
for i in range(10000):
    print(f"Line {i}: " + "x" * 100)
"""

    # Set a small truncation limit for testing
    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CODE_INTERPRETER_MAX_OUTPUT_LENGTH",
        5000,
    ):
        python_tool = PythonTool(tool_id=1, emitter=mock_emitter)
        override_kwargs = PythonToolOverrideKwargs(chat_files=[])
        result = python_tool.run(
            turn_index=0,
            override_kwargs=override_kwargs,
            code=code,
        )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify output was truncated
    assert len(execution_result.stdout) <= 5000 + 200  # Allow for truncation message
    assert "output truncated" in execution_result.stdout
    assert "characters omitted" in execution_result.stdout


def test_python_execution_multiple_files(
    python_tool: PythonTool,
    mock_emitter: Emitter,
    db_session: Session,  # Needed to initialize DB engine for file_store
) -> None:
    """Test generation of multiple files."""
    code = """
# Create multiple files
with open('file1.txt', 'w') as f:
    f.write('Content of file 1')

with open('file2.txt', 'w') as f:
    f.write('Content of file 2')

with open('file3.txt', 'w') as f:
    f.write('Content of file 3')

print("Created 3 files")
"""

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == 0
    assert "Created 3 files" in execution_result.stdout
    assert len(execution_result.generated_files) == 3

    # Verify all files have unique IDs and proper metadata
    file_ids_result = [
        f.file_link.split("/")[-1] for f in execution_result.generated_files
    ]
    assert len(set(file_ids_result)) == 3  # All unique

    # Verify filenames
    filenames = [f.filename for f in execution_result.generated_files]
    assert "file1.txt" in filenames
    assert "file2.txt" in filenames
    assert "file3.txt" in filenames

    # Verify we can read all files back from the file store
    file_store = get_default_file_store()

    # Create a mapping of filename to generated file for easier verification
    files_by_name = {f.filename: f for f in execution_result.generated_files}

    # Verify each expected file
    for i in range(1, 4):
        filename = f"file{i}.txt"
        assert filename in files_by_name, f"Expected file {filename} not found"

        generated_file = files_by_name[filename]
        file_id = generated_file.file_link.split("/")[-1]
        file_io = file_store.read_file(file_id)
        file_content = file_io.read()
        expected_content = f"Content of file {i}".encode()
        assert (
            expected_content in file_content
        ), f"Expected content not found in {filename}"


def test_python_execution_client_error_handling(
    mock_emitter: Emitter,
) -> None:
    """Test error handling when Code Interpreter service fails."""
    from unittest.mock import patch

    code = 'print("Test")'

    if not CODE_INTERPRETER_BASE_URL:
        pytest.skip("CODE_INTERPRETER_BASE_URL not configured")

    # Mock the CodeInterpreterClient to raise an exception
    with patch(
        "onyx.tools.tool_implementations.python.python_tool.CodeInterpreterClient"
    ) as mock_client_class:
        mock_client = Mock()
        mock_client.execute.side_effect = Exception("Service unavailable")
        mock_client_class.return_value = mock_client

        python_tool = PythonTool(tool_id=1, emitter=mock_emitter)
        override_kwargs = PythonToolOverrideKwargs(chat_files=[])
        result = python_tool.run(
            turn_index=0,
            override_kwargs=override_kwargs,
            code=code,
        )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify error result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == -1
    error_msg = execution_result.error or ""
    assert (
        "Service unavailable" in execution_result.stderr
        or "Service unavailable" in error_msg
    )
    assert not execution_result.timed_out
    assert len(execution_result.generated_files) == 0


def test_python_execution_with_excel_file(
    python_tool: PythonTool,
    mock_emitter: Emitter,
    db_session: Session,  # Needed to initialize DB engine for file_store
) -> None:
    """Test Excel file generation with financial data."""
    code = """
import pandas as pd

# Create financial sample data
data = {
    'Segment': ['Government', 'Government', 'Midmarket', 'Midmarket', 'Enterprise'],
    'Country': ['Canada', 'Germany', 'France', 'Germany', 'Canada'],
    'Product': ['Carretera', 'Carretera', 'Carretera', 'Carretera', 'Amarilla'],
    'Units Sold': [1618.5, 1321, 2178, 888, 2470],
    'Manufacturing Price': [3, 3, 3, 3, 260],
    'Sale Price': [20, 20, 20, 20, 300],
    'Gross Sales': [32370, 26420, 43560, 17760, 741000],
    'Discounts': [0, 0, 0, 0, 0],
    'Sales': [32370, 26420, 43560, 17760, 741000],
    'COGS': [16850, 13940, 22800, 9390, 642000],
    'Profit': [15520, 12480, 20760, 8370, 99000],
    'Month': ['January', 'January', 'June', 'April', 'September']
}

# Create DataFrame
df = pd.DataFrame(data)

# Write to Excel
df.to_excel('financial_report.xlsx', index=False, sheet_name='Financial Data')

print(f"Excel file created with {len(df)} rows")
"""

    # Execute code
    override_kwargs = PythonToolOverrideKwargs(chat_files=[])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == 0
    assert "Excel file created with 5 rows" in execution_result.stdout
    assert len(execution_result.generated_files) == 1

    # Verify file metadata
    generated_file = execution_result.generated_files[0]
    assert generated_file.filename == "financial_report.xlsx"
    assert ".xlsx" in generated_file.filename

    # Extract file_id from file_link
    file_id = generated_file.file_link.split("/")[-1]

    # Verify we can read the file back from the file store
    file_store = get_default_file_store()
    file_io = file_store.read_file(file_id)
    file_content = file_io.read()

    # Verify the file is a valid Excel file (check ZIP magic bytes - xlsx is a ZIP archive)
    # ZIP magic bytes: 50 4B 03 04
    assert file_content[:4] == b"PK\x03\x04"
    assert len(file_content) > 1000  # Excel file should be substantial

    # Verify we can parse the Excel file with openpyxl directly
    file_io_obj = io.BytesIO(file_content)
    workbook = load_workbook(file_io_obj)
    sheet = workbook["Financial Data"]

    # Verify data structure - get headers from first row
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = list(first_row) if first_row else []
    expected_columns = [
        "Segment",
        "Country",
        "Product",
        "Units Sold",
        "Manufacturing Price",
        "Sale Price",
        "Gross Sales",
        "Discounts",
        "Sales",
        "COGS",
        "Profit",
        "Month",
    ]
    assert headers == expected_columns

    # Verify row count (excluding header)
    assert sheet.max_row == 6  # 1 header + 5 data rows

    # Read data rows
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(row)

    assert len(rows) == 5

    # Verify some sample data
    segments = [row[0] for row in rows]
    countries = [row[1] for row in rows]

    def _to_float(value: str | float | datetime | None) -> float:
        """Convert value to float, handling datetime objects."""
        if value is None:
            return 0.0
        if isinstance(value, datetime):
            return value.timestamp()
        return float(value)

    units_sold = [_to_float(row[3]) for row in rows]
    profits = [_to_float(row[10]) for row in rows]

    assert "Government" in segments
    assert "Canada" in countries
    assert sum(units_sold) > 8000  # Total units sold
    assert sum(profits) > 155000  # Total profit


def test_python_execution_with_excel_file_input(
    python_tool: PythonTool,
    mock_emitter: Emitter,
    db_session: Session,  # Needed to initialize DB engine for file_store
) -> None:
    """Test processing an uploaded Excel file - reading and analyzing it."""
    import os

    test_file_path = os.path.join(
        os.path.dirname(__file__), "data", "financial-sample.xlsx"
    )

    with open(test_file_path, "rb") as f:
        file_content = f.read()

    # Create InMemoryChatFile with the Excel file
    in_memory_chat_file = InMemoryChatFile(
        file_id="test-financial-sample",
        content=file_content,
        file_type=ChatFileType.DOC,
        filename="financial-sample.xlsx",
    )

    # Convert InMemoryChatFile to ChatFile for the tool
    chat_file = ChatFile(
        filename=in_memory_chat_file.filename or "financial-sample.xlsx",
        content=in_memory_chat_file.content,
    )

    # Code to analyze the uploaded Excel file
    code = """
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Read the uploaded Excel file using openpyxl directly
workbook = load_workbook('financial-sample.xlsx')
sheet = workbook.active

# Convert to pandas DataFrame
data = []
headers = [cell.value for cell in sheet[1]]
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)

df = pd.DataFrame(data, columns=headers)

print(f"Loaded Excel file with {len(df)} rows and {len(df.columns)} columns")
print(f"\\nColumns: {', '.join(df.columns.tolist())}")

# Perform analysis
print(f"\\n=== Analysis ===")

# Group by segment and calculate total sales and profit
segment_summary = df.groupby('Segment').agg({
    ' Sales': 'sum',
    'Profit': 'sum',
    'Units Sold': 'sum'
}).round(2)

print(f"\\nSales by Segment:")
print(segment_summary)

# Find top 5 products by profit
top_products = df.groupby('Product')['Profit'].sum().sort_values(ascending=False).head(5)
print(f"\\nTop 5 Products by Profit:")
print(top_products)

# Calculate profit margin
total_sales = df[' Sales'].sum()
total_profit = df['Profit'].sum()
profit_margin = (total_profit / total_sales * 100) if total_sales > 0 else 0
print(f"\\nOverall Profit Margin: {profit_margin:.2f}%")

# Create a visualization
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

# Sales by Segment
segment_summary[' Sales'].plot(kind='bar', ax=ax1, color='steelblue')
ax1.set_title('Total Sales by Segment')
ax1.set_xlabel('Segment')
ax1.set_ylabel('Sales ($)')
ax1.tick_params(axis='x', rotation=45)

# Top 5 Products by Profit
top_products.plot(kind='barh', ax=ax2, color='seagreen')
ax2.set_title('Top 5 Products by Profit')
ax2.set_xlabel('Profit ($)')
ax2.set_ylabel('Product')

plt.tight_layout()
plt.savefig('financial_analysis.png', dpi=100, bbox_inches='tight')
print(f"\\nVisualization saved as financial_analysis.png")

# Create summary report Excel file
summary_data = {
    'Metric': ['Total Sales', 'Total Profit', 'Profit Margin %', 'Total Units Sold', 'Number of Records'],
    'Value': [
        f"${total_sales:,.2f}",
        f"${total_profit:,.2f}",
        f"{profit_margin:.2f}%",
        f"{df['Units Sold'].sum():,.0f}",
        len(df)
    ]
}
summary_df = pd.DataFrame(summary_data)

with pd.ExcelWriter('financial_summary.xlsx') as writer:
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    segment_summary.to_excel(writer, sheet_name='By Segment')

print(f"Summary report saved as financial_summary.xlsx")
"""

    # Execute code with the uploaded file
    override_kwargs = PythonToolOverrideKwargs(chat_files=[chat_file])
    result = python_tool.run(
        turn_index=0,
        override_kwargs=override_kwargs,
        code=code,
    )

    # Parse result
    adapter = TypeAdapter(LlmPythonExecutionResult)
    execution_result = adapter.validate_json(result.llm_facing_response)

    # Verify result
    assert isinstance(execution_result, LlmPythonExecutionResult)
    assert execution_result.exit_code == 0
    assert "Loaded Excel file" in execution_result.stdout
    assert "Analysis" in execution_result.stdout
    assert "Sales by Segment" in execution_result.stdout
    assert "Top 5 Products by Profit" in execution_result.stdout
    assert "Profit Margin" in execution_result.stdout

    # Should generate 2 files: PNG visualization and Excel summary
    assert len(execution_result.generated_files) == 2

    # Verify generated files
    filenames = [f.filename for f in execution_result.generated_files]
    assert "financial_analysis.png" in filenames
    assert "financial_summary.xlsx" in filenames

    # Verify we can read and validate the generated files
    file_store = get_default_file_store()

    # Check the PNG file
    png_file = next(
        f
        for f in execution_result.generated_files
        if f.filename == "financial_analysis.png"
    )
    png_file_id = png_file.file_link.split("/")[-1]
    png_io = file_store.read_file(png_file_id)
    png_content = png_io.read()
    assert png_content[:8] == b"\x89PNG\r\n\x1a\n"  # PNG magic bytes
    assert len(png_content) > 5000  # Should be substantial

    # Check the Excel summary file
    xlsx_file = next(
        f
        for f in execution_result.generated_files
        if f.filename == "financial_summary.xlsx"
    )
    xlsx_file_id = xlsx_file.file_link.split("/")[-1]
    xlsx_io = file_store.read_file(xlsx_file_id)
    xlsx_content = xlsx_io.read()
    assert xlsx_content[:4] == b"PK\x03\x04"  # ZIP/Excel magic bytes

    # Parse and verify the summary Excel file using openpyxl directly
    xlsx_io_obj = io.BytesIO(xlsx_content)
    workbook = load_workbook(xlsx_io_obj)
    sheet = workbook["Summary"]

    # Read headers from first row
    first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = list(first_row) if first_row else []
    assert "Metric" in headers
    assert "Value" in headers

    # Read all rows and extract metrics
    metrics = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Metric column
            metrics.append(row[0])

    assert "Total Sales" in metrics
    assert "Total Profit" in metrics
    assert "Profit Margin %" in metrics


if __name__ == "__main__":
    # Run with: python -m pytest tests/external_dependency_unit/tools/test_python_tool.py -v
    pytest.main([__file__, "-v"])
